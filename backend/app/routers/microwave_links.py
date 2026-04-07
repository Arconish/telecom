import io

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import select, or_, func
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.models.microwave_link import MicrowaveLink
from app.models.user import User
from app.routers.auth import require_admin
from app.schemas.microwave_link import (
    MicrowaveLinkCreate,
    MicrowaveLinkOut,
    MicrowaveLinkUpdate,
)

router = APIRouter(prefix="/microwave-links", tags=["microwave-links"])


ALLOWED_SORT_FIELDS = {
    "id": MicrowaveLink.id,
    "ne_id": MicrowaveLink.ne_id,
    "fe_id": MicrowaveLink.fe_id,
    "link_id": MicrowaveLink.link_id,
    "management_ip": MicrowaveLink.management_ip,
    "web_protocol": MicrowaveLink.web_protocol,
    "link_class": MicrowaveLink.link_class,
    "is_active": MicrowaveLink.is_active,
    "vendor": MicrowaveLink.vendor,
    "model": MicrowaveLink.model,
    "type": MicrowaveLink.type,
    "status": MicrowaveLink.status,
    "created_at": MicrowaveLink.created_at,
    "updated_at": MicrowaveLink.updated_at,
}

EXPORT_HEADERS = [
    "ne_id",
    "fe_id",
    "link_id",
    "management_ip",
    "web_protocol",
    "link_class",
    "is_active",
    "vendor",
    "model",
    "type",
    "status",
]

REQUIRED_HEADERS = {"link_id"}

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1E40AF")
REQUIRED_HEADER_FILL = PatternFill(fill_type="solid", fgColor="DC2626")
HEADER_FONT = Font(color="FFFFFF", bold=True)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)


def apply_common_filters(
    base_query,
    search: str | None = None,
    status: str | None = None,
    is_active: bool | None = None,
    vendor: str | None = None,
):
    if search:
        like_term = f"%{search}%"
        base_query = base_query.where(
            or_(
                MicrowaveLink.ne_id.ilike(like_term),
                MicrowaveLink.fe_id.ilike(like_term),
                MicrowaveLink.link_id.ilike(like_term),
                MicrowaveLink.management_ip.ilike(like_term),
                MicrowaveLink.vendor.ilike(like_term),
                MicrowaveLink.model.ilike(like_term),
                MicrowaveLink.type.ilike(like_term),
                MicrowaveLink.status.ilike(like_term),
                MicrowaveLink.link_class.ilike(like_term),
            )
        )

    if status:
        base_query = base_query.where(MicrowaveLink.status == status)

    if is_active is not None:
        base_query = base_query.where(MicrowaveLink.is_active == is_active)

    if vendor:
        base_query = base_query.where(MicrowaveLink.vendor == vendor)

    return base_query


def build_paginated_response(items, total: int, page: int, page_size: int):
    total_pages = (total + page_size - 1) // page_size

    return {
        "items": [
            {
                "id": row.id,
                "ne_id": row.ne_id,
                "fe_id": row.fe_id,
                "link_id": row.link_id,
                "management_ip": row.management_ip,
                "web_protocol": row.web_protocol,
                "link_class": row.link_class,
                "is_active": row.is_active,
                "vendor": row.vendor,
                "model": row.model,
                "type": row.type,
                "status": row.status,
                "created_at": row.created_at,
                "updated_at": row.updated_at,
            }
            for row in items
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
    }


def parse_bool(value):
    if isinstance(value, bool):
        return value

    if value is None or value == "":
        return None

    normalized = str(value).strip().lower()

    if normalized in {"true", "1", "yes", "y", "active"}:
        return True
    if normalized in {"false", "0", "no", "n", "down", "inactive"}:
        return False

    raise ValueError("Invalid boolean value")


def style_header_row(ws, headers):
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_index)
        cell.value = header
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
        cell.fill = REQUIRED_HEADER_FILL if header in REQUIRED_HEADERS else HEADER_FILL


def auto_fit_columns(ws):
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))

        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 28)


def add_data_validations(ws, start_row=2, end_row=1000):
    protocol_validation = DataValidation(
        type="list",
        formula1='"http,https"',
        allow_blank=True,
    )
    status_validation = DataValidation(
        type="list",
        formula1='"On Air,Planned,Down"',
        allow_blank=True,
    )
    active_validation = DataValidation(
        type="list",
        formula1='"TRUE,FALSE"',
        allow_blank=True,
    )

    ws.add_data_validation(protocol_validation)
    ws.add_data_validation(status_validation)
    ws.add_data_validation(active_validation)

    header_to_col = {header: idx + 1 for idx, header in enumerate(EXPORT_HEADERS)}

    protocol_col = ws.cell(row=1, column=header_to_col["web_protocol"]).column_letter
    status_col = ws.cell(row=1, column=header_to_col["status"]).column_letter
    active_col = ws.cell(row=1, column=header_to_col["is_active"]).column_letter

    protocol_validation.add(f"{protocol_col}{start_row}:{protocol_col}{end_row}")
    status_validation.add(f"{status_col}{start_row}:{status_col}{end_row}")
    active_validation.add(f"{active_col}{start_row}:{active_col}{end_row}")


def build_instruction_sheet(wb):
    ws = wb.create_sheet(title="Instructions")

    rows = [
        ["Microwave Links Excel Template"],
        [""],
        ["How to use"],
        ["1. Fill data in the 'Microwave Links' sheet."],
        ["2. Do not rename the required headers."],
        ["3. 'link_id' is required and must be unique for new records."],
        ["4. If link_id already exists, import will update that row."],
        ["5. Allowed web_protocol values: http, https"],
        ["6. Allowed status values: On Air, Planned, Down"],
        ["7. Allowed is_active values: TRUE, FALSE"],
        [""],
        ["Required columns"],
        ["link_id"],
        [""],
        ["Column descriptions"],
        ["ne_id", "Near End ID"],
        ["fe_id", "Far End ID"],
        ["link_id", "Unique microwave link ID"],
        ["management_ip", "Management IP address"],
        ["web_protocol", "http or https"],
        ["link_class", "Link class / category"],
        ["is_active", "TRUE or FALSE"],
        ["vendor", "Vendor name"],
        ["model", "Model name"],
        ["type", "Link type"],
        ["status", "On Air / Planned / Down"],
    ]

    for row in rows:
        ws.append(row)

    ws["A1"].font = Font(size=14, bold=True)
    ws["A3"].font = Font(bold=True)
    ws["A11"].font = Font(bold=True)
    ws["A14"].font = Font(bold=True)

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 40


@router.post("/", response_model=MicrowaveLinkOut)
def create_microwave_link(
    payload: MicrowaveLinkCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    existing = db.execute(
        select(MicrowaveLink).where(MicrowaveLink.link_id == payload.link_id)
    ).scalar_one_or_none()

    if existing:
        raise HTTPException(status_code=400, detail="Link ID already exists")

    row = MicrowaveLink(**payload.model_dump())
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@router.get("/")
def list_microwave_links(
    db: Session = Depends(get_db),
    search: str | None = Query(default=None),
    status: str | None = Query(default=None),
    is_active: bool | None = Query(default=None),
    vendor: str | None = Query(default=None),
    sort_by: str = Query(default="link_id"),
    sort_order: str = Query(default="asc"),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=10, ge=1, le=200),
):
    if sort_by not in ALLOWED_SORT_FIELDS:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid sort_by. Allowed values: {', '.join(ALLOWED_SORT_FIELDS.keys())}",
        )

    if sort_order not in {"asc", "desc"}:
        raise HTTPException(
            status_code=400,
            detail="Invalid sort_order. Allowed values: asc, desc",
        )

    base_query = select(MicrowaveLink)
    base_query = apply_common_filters(
        base_query=base_query,
        search=search,
        status=status,
        is_active=is_active,
        vendor=vendor,
    )

    count_query = select(func.count()).select_from(base_query.subquery())
    total = db.execute(count_query).scalar() or 0

    offset = (page - 1) * page_size
    sort_column = ALLOWED_SORT_FIELDS[sort_by]
    order_clause = sort_column.asc() if sort_order == "asc" else sort_column.desc()

    paged_query = base_query.order_by(order_clause).offset(offset).limit(page_size)
    items = db.execute(paged_query).scalars().all()

    return build_paginated_response(items, total, page, page_size)


@router.get("/status/view")
def list_microwave_link_status(
    db: Session = Depends(get_db),
    search: str | None = Query(default=None),
    status: str | None = Query(default=None),
    is_active: bool | None = Query(default=None),
    vendor: str | None = Query(default=None),
    sort_by: str = Query(default="link_id"),
    sort_order: str = Query(default="asc"),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=10, ge=1, le=200),
):
    if sort_by not in ALLOWED_SORT_FIELDS:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid sort_by. Allowed values: {', '.join(ALLOWED_SORT_FIELDS.keys())}",
        )

    if sort_order not in {"asc", "desc"}:
        raise HTTPException(
            status_code=400,
            detail="Invalid sort_order. Allowed values: asc, desc",
        )

    base_query = select(MicrowaveLink)
    base_query = apply_common_filters(
        base_query=base_query,
        search=search,
        status=status,
        is_active=is_active,
        vendor=vendor,
    )

    count_query = select(func.count()).select_from(base_query.subquery())
    total = db.execute(count_query).scalar() or 0

    offset = (page - 1) * page_size
    sort_column = ALLOWED_SORT_FIELDS[sort_by]
    order_clause = sort_column.asc() if sort_order == "asc" else sort_column.desc()

    paged_query = base_query.order_by(order_clause).offset(offset).limit(page_size)
    items = db.execute(paged_query).scalars().all()

    return build_paginated_response(items, total, page, page_size)


@router.get("/status/summary")
def microwave_link_status_summary(
    db: Session = Depends(get_db),
):
    total_links = db.execute(
        select(func.count()).select_from(MicrowaveLink)
    ).scalar() or 0

    active_links = db.execute(
        select(func.count()).select_from(MicrowaveLink).where(MicrowaveLink.is_active == True)
    ).scalar() or 0

    inactive_links = total_links - active_links

    status_rows = db.execute(
        select(
            MicrowaveLink.status,
            func.count().label("count")
        ).group_by(MicrowaveLink.status)
    ).all()

    vendor_rows = db.execute(
        select(
            MicrowaveLink.vendor,
            func.count().label("count")
        ).group_by(MicrowaveLink.vendor)
    ).all()

    status_counts = {
        (status if status else "Unknown"): count
        for status, count in status_rows
    }

    vendor_counts = {
        (vendor if vendor else "Unknown"): count
        for vendor, count in vendor_rows
    }

    return {
        "total_links": total_links,
        "active_links": active_links,
        "inactive_links": inactive_links,
        "status_counts": status_counts,
        "vendor_counts": vendor_counts,
    }


@router.get("/export/excel")
def export_microwave_links_excel(
    db: Session = Depends(get_db),
    search: str | None = Query(default=None),
    status: str | None = Query(default=None),
    is_active: bool | None = Query(default=None),
    vendor: str | None = Query(default=None),
    sort_by: str = Query(default="link_id"),
    sort_order: str = Query(default="asc"),
):
    if sort_by not in ALLOWED_SORT_FIELDS:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid sort_by. Allowed values: {', '.join(ALLOWED_SORT_FIELDS.keys())}",
        )

    if sort_order not in {"asc", "desc"}:
        raise HTTPException(
            status_code=400,
            detail="Invalid sort_order. Allowed values: asc, desc",
        )

    base_query = select(MicrowaveLink)
    base_query = apply_common_filters(
        base_query=base_query,
        search=search,
        status=status,
        is_active=is_active,
        vendor=vendor,
    )

    sort_column = ALLOWED_SORT_FIELDS[sort_by]
    order_clause = sort_column.asc() if sort_order == "asc" else sort_column.desc()

    items = db.execute(base_query.order_by(order_clause)).scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Microwave Links"

    style_header_row(ws, EXPORT_HEADERS)

    for row in items:
        ws.append([
            row.ne_id or "",
            row.fe_id or "",
            row.link_id or "",
            row.management_ip or "",
            row.web_protocol or "",
            row.link_class or "",
            row.is_active,
            row.vendor or "",
            row.model or "",
            row.type or "",
            row.status or "",
        ])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:K{max(ws.max_row, 2)}"

    add_data_validations(ws, start_row=2, end_row=max(ws.max_row + 200, 1000))
    auto_fit_columns(ws)
    build_instruction_sheet(wb)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=microwave_links_template.xlsx"
        },
    )


@router.post("/import/excel")
def import_microwave_links_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    allowed_extensions = (".xlsx", ".xlsm")
    if not file.filename.lower().endswith(allowed_extensions):
        raise HTTPException(
            status_code=400,
            detail="Only Excel files (.xlsx, .xlsm) are allowed",
        )

    file_bytes = file.file.read()
    workbook = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)

    if "Microwave Links" in workbook.sheetnames:
        worksheet = workbook["Microwave Links"]
    else:
        worksheet = workbook.active

    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel file is empty")

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    missing_columns = [col for col in EXPORT_HEADERS if col not in headers]

    if missing_columns:
        raise HTTPException(
            status_code=400,
            detail=f"Missing required Excel columns: {', '.join(missing_columns)}",
        )

    header_index = {header: idx for idx, header in enumerate(headers)}

    created = 0
    updated = 0
    errors = []

    for excel_row_number, row_values in enumerate(rows[1:], start=2):
        def get_cell(name):
            idx = header_index[name]
            if idx >= len(row_values):
                return None
            return row_values[idx]

        link_id = str(get_cell("link_id") or "").strip()

        if not link_id:
            if all(cell is None or str(cell).strip() == "" for cell in row_values):
                continue
            errors.append(f"Row {excel_row_number}: link_id is required")
            continue

        try:
            is_active = parse_bool(get_cell("is_active"))
        except ValueError:
            errors.append(f"Row {excel_row_number}: invalid is_active value")
            continue

        payload = {
            "ne_id": str(get_cell("ne_id") or "").strip(),
            "fe_id": str(get_cell("fe_id") or "").strip(),
            "link_id": link_id,
            "management_ip": str(get_cell("management_ip") or "").strip(),
            "web_protocol": str(get_cell("web_protocol") or "http").strip() or "http",
            "link_class": str(get_cell("link_class") or "").strip(),
            "is_active": True if is_active is None else is_active,
            "vendor": str(get_cell("vendor") or "").strip(),
            "model": str(get_cell("model") or "").strip(),
            "type": str(get_cell("type") or "").strip(),
            "status": str(get_cell("status") or "").strip(),
        }

        existing = db.execute(
            select(MicrowaveLink).where(MicrowaveLink.link_id == link_id)
        ).scalar_one_or_none()

        if existing:
            for field, value in payload.items():
                setattr(existing, field, value)
            updated += 1
        else:
            db.add(MicrowaveLink(**payload))
            created += 1

    db.commit()

    return {
        "message": "Excel import completed",
        "created": created,
        "updated": updated,
        "errors": errors,
    }


@router.get("/{row_id}", response_model=MicrowaveLinkOut)
def get_microwave_link(
    row_id: int,
    db: Session = Depends(get_db),
):
    row = db.execute(
        select(MicrowaveLink).where(MicrowaveLink.id == row_id)
    ).scalar_one_or_none()

    if not row:
        raise HTTPException(status_code=404, detail="Microwave link not found")

    return row


@router.put("/{row_id}", response_model=MicrowaveLinkOut)
def update_microwave_link(
    row_id: int,
    payload: MicrowaveLinkUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    row = db.execute(
        select(MicrowaveLink).where(MicrowaveLink.id == row_id)
    ).scalar_one_or_none()

    if not row:
        raise HTTPException(status_code=404, detail="Microwave link not found")

    update_data = payload.model_dump(exclude_unset=True)

    if "link_id" in update_data and update_data["link_id"] != row.link_id:
        existing = db.execute(
            select(MicrowaveLink).where(MicrowaveLink.link_id == update_data["link_id"])
        ).scalar_one_or_none()

        if existing:
            raise HTTPException(status_code=400, detail="Link ID already exists")

    for field, value in update_data.items():
        setattr(row, field, value)

    db.commit()
    db.refresh(row)
    return row


@router.delete("/{row_id}")
def delete_microwave_link(
    row_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    row = db.execute(
        select(MicrowaveLink).where(MicrowaveLink.id == row_id)
    ).scalar_one_or_none()

    if not row:
        raise HTTPException(status_code=404, detail="Microwave link not found")

    db.delete(row)
    db.commit()

    return {"message": "Microwave link deleted successfully"}